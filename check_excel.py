from openpyxl import load_workbook

# 加载Excel文件
workbook = load_workbook('/Users/rayz/Downloads/yuki-cidoc-proj/文物文化特征单元数据结构.xlsx')

# 打印所有工作表名称
print('所有工作表名称:', workbook.sheetnames)

# 选择'structure v2'工作表
sheet = workbook['structure v2']

# 打印前几行
column_names = [cell.value for cell in sheet[1]]
print('\n"structure v2" 工作表的列名:')
for i, name in enumerate(column_names):
    print(f'{i+1}. {name}')